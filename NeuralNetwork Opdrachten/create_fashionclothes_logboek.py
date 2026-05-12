import tensorflow as tf
from tensorflow.keras.datasets import fashion_mnist
import openpyxl
from openpyxl.styles import Font, PatternFill
import random

# Load data
(train_data, train_labels), (test_data, test_labels) = fashion_mnist.load_data()
train_data_norm = train_data / 255.0
test_data_norm = test_data / 255.0

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "FashionClothes Testen"

# Headers
headers = ["Test", "Layer1", "Layer2", "Optimizer", "Learning Rate", "Epochs", "Batch Size", 
           "Train Acc", "Val Acc", "Train Loss", "Val Loss"]
ws.append(headers)

# Format header
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Hyperparameter combinations for 30 tests
test_configs = [
    # Layer configs: (layer1_size, layer2_size, optimizer, lr, epochs, batch_size)
    (32, 16, "adam", 0.001, 5, 32),
    (32, 16, "adam", 0.001, 10, 32),
    (64, 32, "adam", 0.001, 5, 32),
    (64, 32, "adam", 0.001, 10, 32),
    (128, 64, "adam", 0.001, 5, 32),
    (128, 64, "adam", 0.001, 10, 32),
    (32, 16, "adam", 0.01, 5, 32),
    (32, 16, "adam", 0.01, 10, 32),
    (64, 32, "adam", 0.01, 5, 32),
    (64, 32, "adam", 0.01, 10, 32),
    (128, 64, "adam", 0.01, 5, 32),
    (128, 64, "adam", 0.01, 10, 32),
    (32, 16, "adam", 0.0001, 10, 32),
    (64, 32, "adam", 0.0001, 10, 32),
    (128, 64, "adam", 0.0001, 10, 32),
    (32, 16, "rmsprop", 0.001, 10, 32),
    (64, 32, "rmsprop", 0.001, 10, 32),
    (128, 64, "rmsprop", 0.001, 10, 32),
    (32, 16, "sgd", 0.01, 10, 32),
    (64, 32, "sgd", 0.01, 10, 32),
    (128, 64, "sgd", 0.01, 10, 32),
    (256, 128, "adam", 0.001, 5, 32),
    (256, 128, "adam", 0.001, 10, 32),
    (256, 128, "adam", 0.01, 10, 32),
    (32, 16, "adam", 0.001, 10, 64),
    (64, 32, "adam", 0.001, 10, 64),
    (128, 64, "adam", 0.001, 10, 64),
    (32, 16, "adam", 0.001, 10, 16),
    (64, 32, "adam", 0.001, 10, 16),
    (128, 64, "adam", 0.001, 10, 16),
]

# Run 30 tests
for i, config in enumerate(test_configs, 1):
    layer1, layer2, optimizer, lr, epochs, batch_size = config
    
    print(f"Test {i}/30: layers({layer1},{layer2}), {optimizer} lr={lr}, epochs={epochs}, batch={batch_size}")
    
    # Build model
    model = tf.keras.Sequential([
        tf.keras.layers.Flatten(input_shape=train_data[0].shape),
        tf.keras.layers.Dense(layer1, activation=tf.keras.activations.relu),
        tf.keras.layers.Dense(layer2, activation=tf.keras.activations.relu),
        tf.keras.layers.Dense(10, activation=tf.keras.activations.softmax)
    ])
    
    # Compile
    if optimizer == "adam":
        opt = tf.keras.optimizers.Adam(lr)
    elif optimizer == "rmsprop":
        opt = tf.keras.optimizers.RMSprop(lr)
    else:  # sgd
        opt = tf.keras.optimizers.SGD(lr)
    
    model.compile(loss=tf.keras.losses.SparseCategoricalCrossentropy(),
                  optimizer=opt,
                  metrics=["Accuracy"])
    
    # Train
    history = model.fit(train_data_norm, train_labels,
                       epochs=epochs,
                       batch_size=batch_size,
                       validation_data=(test_data_norm, test_labels),
                       verbose=0)
    
    # Get metrics
    train_acc = history.history['Accuracy'][-1]
    val_acc = history.history['val_Accuracy'][-1]
    train_loss = history.history['loss'][-1]
    val_loss = history.history['val_loss'][-1]
    
    # Add to spreadsheet
    ws.append([i, layer1, layer2, optimizer, lr, epochs, batch_size, 
               f"{train_acc:.4f}", f"{val_acc:.4f}", f"{train_loss:.4f}", f"{val_loss:.4f}"])

# Save workbook
wb.save("Logboek_FashionClothes.xlsx")
print("\n✅ Logboek opgeslagen: Logboek_FashionClothes.xlsx")
print("30 testen afgerond!")
